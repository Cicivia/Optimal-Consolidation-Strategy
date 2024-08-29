import openpyxl
import pandas as pd
import loadmsg
from collections import OrderedDict

def dijkstra(path, start_city, end_city):
    # 利用dijkstra算法计算出每个城市到起始城市的最短路径
    dis = {}  # 保存每个城市到起始城市的最短距离
    path_dict = {}  # 保存每个城市的最短路径的前驱结点
    # 初始化dis和path_dict字典
    for city in set(path['起始城市']).union(set(path['目的城市'])):
        dis[city] = float('inf')  # 初始距离为无穷大
        path_dict[city] = None  # 初始化前驱结点为None
    dis[start_city] = 0  # 起始城市到自己的距离为0
    visited = set()  # 记录已经访问过的城市
    while len(visited) < len(dis):
        # 选取当前距离起点最近的未访问城市
        cur_city = min(dis.items(), key=lambda x: x[1] if x[0] not in visited else float('inf'))[0]
        visited.add(cur_city)
        # 更新与当前城市相邻的城市的距离
        for _, row in path.loc[path['起始城市'] == cur_city].iterrows():
            neighbor_city = row['目的城市']
            new_dis = dis[cur_city] + int(row['距离'])  # 计算从起点到当前路径终点的距离， 即从当前城市到起点城市加上当前路径的长度
            if new_dis < dis[neighbor_city]:  # 如果从起点城市经过当前城市到达终点城市的城市比已知最短距离更小，则更新该城市到起点城市的距离和前驱结点
                dis[neighbor_city] = new_dis
                path_dict[neighbor_city] = cur_city
    # 从终点城市逆推出起点城市到终点城市的最短路径
    path_list = [end_city]  # 保存路径信息的列表
    distance = dis[end_city]
    while path_dict[end_city] != start_city:
        path_list.append(path_dict[end_city])
        end_city = path_dict[end_city]
    path_list.append(start_city)
    path_list.reverse()  # 将列表中的元素顺序反转，得到起点和终点的路径信息
    return path_list, distance

def sort_order():   # 修改订单数据，返回处理后的订单数据(只根据线路信息计算起始城市之间的距离，修改订单)
    excel = openpyxl.load_workbook('集货策略赛题数据.xlsx')
    sheet1 = excel['赛题数据-可用路线']
    distances = {}  # 存储城市之间的距离信息
    i = 2
    while sheet1['A' + str(i)].value is not None:
        start_city = sheet1['B' + str(i)].value
        end_city = sheet1['C' + str(i)].value
        distance = sheet1['D' + str(i)].value
        distances[(start_city, end_city)] = distance
        i += 1
    # 获取订单信息
    orders = loadmsg.load('订单信息')
    grouped = orders.groupby('起始城市')  # 将起始城市相同的订单放在一起
    start_cities = list(grouped.groups.keys())
    result = pd.DataFrame(columns=['订单编号', '起始城市', '目的城市'])  # 保存最后结果
    max = 500  # 规定临近城市之间的距离为max
    for i, start_city in enumerate(start_cities):
        orders_start = grouped.get_group(start_city)
        result = pd.concat([result, orders_start]).drop_duplicates(keep='first')
        for j in range(i + 1, len(start_cities)):
            end_city = start_cities[j]
            if (start_city, end_city) in distances:
                distance = distances[start_city, end_city]
                if (int(distance) < max):
                    orders_end = grouped.get_group(end_city)
                    orders_around = pd.concat([orders_start, orders_end])
                    result = pd.concat([result, orders_around]).drop_duplicates(keep='first')
            elif (end_city, start_city) in distances:
                distance = distances[end_city, start_city]
                if (int(distance) < max):
                    orders_end = grouped.get_group(end_city)
                    orders_around = pd.concat([orders_start, orders_end])
                    result = pd.concat([result, orders_around]).drop_duplicates(keep='first')
    return result

def sort_order2():  # 修改订单数据，返回处理后的订单数据(只根据线路信息计算起始城市之间的距离，修改订单)
    excel = openpyxl.load_workbook('集货策略赛题数据.xlsx')
    sheet1 = excel['赛题数据-可用路线']
    distances = {}  # 存储城市之间的距离信息
    i = 2
    while sheet1['A' + str(i)].value is not None:
        start_city = sheet1['B' + str(i)].value
        end_city = sheet1['C' + str(i)].value
        distance = sheet1['D' + str(i)].value
        distances[(start_city, end_city)] = distance
        i += 1
    # 获取订单信息
    orders = loadmsg.load('订单信息')
    grouped = orders.groupby('起始城市')  # 将起始城市相同的订单放在一起
    # start_cities = list(set(orders['起始城市']))
    start_cities = list(grouped.groups.keys())
    result = pd.DataFrame(columns=['订单编号', '起始城市', '目的城市'])  # 保存最后结果
    max = 500  # 规定临近城市之间的距离为max
    around = OrderedDict.fromkeys(start_cities)  # 保存键值城市的临近城市
    for i, start_city in enumerate(start_cities):  # 得到around数据，以字典形式存储邻近数据
        around[start_city] = []
        for j in range(i + 1, len(start_cities)):
            end_city = start_cities[j]
            if (start_city, end_city) in distances:
                distance = distances[start_city, end_city]
                if (int(distance) < max):
                    around[start_city].append(end_city)
            elif (end_city, start_city) in distances:
                distance = distances[end_city, start_city]
                if (int(distance) < max):
                    around[start_city].append(end_city)
    sorted_around = dict(sorted(around.items(), key=lambda x: len(x[1])))
    # print(sorted_around)
    # for start_city, end_cities in sorted_around.items():    #根据下一个起始地与当前起始地的距离进行排序的距离进行排序
    #     dist = OrderedDict.fromkeys(end_cities)
    #     sorted_dis = {}
    #     for end_city in end_cities:
    #         if (start_city, end_city) in distances:
    #             dist[end_city] = distances[start_city, end_city]
    #         else:
    #             dist[end_city] = distances[end_city, start_city]
    #     sorted_items = sorted(dist.items(), key=lambda x: x[1], reverse=True)
    #     for item in sorted_items:
    #         sorted_dis[item[0]] = item[1]
    #     sorted_around[start_city] = list(sorted_dis.keys())
    # print(sorted_around)
    for start_city, end_cities in sorted_around.items():
        orders_start = grouped.get_group(start_city)
        if end_cities:
            for end_city in end_cities:
                orders_end = grouped.get_group(end_city)
                orders_around = pd.concat([orders_start, orders_end]).drop_duplicates()
                result = pd.concat([result, orders_around]).drop_duplicates(keep='first')
        else:
            result = pd.concat([result, orders_start]).drop_duplicates(keep='first')
    return result

def sort_orders1(path): #只考虑起始城市,使用dijkstra算法
    orders = loadmsg.load('订单信息')
    grouped = orders.groupby('起始城市')
    start_cities = list(set(orders['起始城市']))
    result = pd.DataFrame(columns=['订单编号', '起始城市', '目的城市'])
    excel = openpyxl.load_workbook('集货策略赛题数据.xlsx')
    sheet1 = excel['赛题数据-可用路线']
    distances = {}  # 存储城市之间的距离信息
    i = 2
    while sheet1['A' + str(i)].value is not None:
        start_city = sheet1['B' + str(i)].value
        end_city = sheet1['C' + str(i)].value
        distance = sheet1['D' + str(i)].value
        distances[(start_city, end_city)] = distance
        i += 1
    max = 500
    around = dict.fromkeys(start_cities)
    for i, start_city in enumerate(start_cities):
        around[start_city] = []
        for j in range(i+1, len(start_cities)):
            end_city = start_cities[j]
            if dijkstra(path, start_city, end_city)[1] < max:
                around[start_city].append(end_city)
    sorted_around = dict(sorted(around.items(), key=lambda x: len(x[1]), reverse=True))
    print(around)
    print(len(around))
    for start_city, end_cities in sorted_around.items():
        orders_start = grouped.get_group(start_city)
        if end_cities:
            for end_city in end_cities:
                orders_end = grouped.get_group(end_city)
                orders_around = pd.concat([orders_start, orders_end]).drop_duplicates()
                result = pd.concat([result, orders_around]).drop_duplicates(keep='first')
        else:
            result = pd.concat([result, orders_start]).drop_duplicates(keep='first')
    return result


def r_sort_orders():
    test1 = sort_order2()
    test1.to_excel("Data_Order1.xlsx", sheet_name='sheet2', index=False)
    df1 = pd.read_excel("Data_Order1.xlsx", sheet_name='sheet2')
    unique_value1 = df1['起始城市'].nunique()
    print(unique_value1)
    unique_count1 = df1['起始城市'].unique()
    startcities = unique_count1
    excel = openpyxl.load_workbook('集货策略赛题数据.xlsx')
    sheet1 = excel['赛题数据-可用路线']
    distances = {}  # 存储城市之间的距离信息
    i = 2
    while sheet1['A' + str(i)].value is not None:
        start_city = sheet1['B' + str(i)].value
        end_city = sheet1['C' + str(i)].value
        distance = sheet1['D' + str(i)].value
        distances[(start_city, end_city)] = distance
        i += 1
    result = pd.DataFrame(columns=['订单编号', '起始城市', '目的城市'])  # 保存最后结果
    start_cities = startcities
    orders = loadmsg.load('订单信息')
    grouped_start = orders.groupby('起始城市')  # 将起始城市相同的订单放在一起
    e_max = 500
    for start_city in start_cities:
        orders_start = grouped_start.get_group(start_city)
        grouped_end = orders_start.groupby('目的城市')
        end_cities = list(grouped_end.groups.keys())
        around = OrderedDict.fromkeys(end_cities)  # 保存键值城市的临近城市
        for i, end_city in enumerate(end_cities):  # 得到around数据，以字典形式存储邻近数据
            around[end_city] = []
            for j in range(i + 1, len(end_cities)):
                end_city1 = end_cities[j]
                if (end_city, end_city1) in distances:
                    distance = distances[end_city, end_city1]
                    if (int(distance) < e_max):
                        around[end_city].append(end_city1)
                elif (end_city1, end_city) in distances:
                    distance = distances[end_city1, end_city]
                    if (int(distance) < e_max):
                        around[end_city].append(end_city1)
        for end_city, end_cities in around.items():
            orders_end = grouped_end.get_group(end_city)
            if end_cities:
                for end_city1 in end_cities:
                    orders_end1 = grouped_end.get_group(end_city1)
                    orders_around = pd.concat([orders_end, orders_end1]).drop_duplicates()
                    result = pd.concat([result, orders_around]).drop_duplicates(keep='first')
            else:
                result = pd.concat([result, orders_end])
    return result





# 测试代码
# test = sort_order()
# test.to_excel("Data_Order.xlsx", index=False)
# df = pd.read_excel("Data_Order.xlsx")
# unique_value = df['起始城市'].nunique()
# print(unique_value)
# unique_count = df['起始城市'].unique()
# print(unique_count)
# 测试代码
# test1 = sort_order2()
# test1.to_excel("Data_Order1.xlsx", sheet_name='sheet2', index=False)
# df1 = pd.read_excel("Data_Order1.xlsx", sheet_name='sheet2')
# unique_value1 = df1['起始城市'].nunique()
# print(unique_value1)
# unique_count1 = df1['起始城市'].unique()
# print(unique_count1)
# print('修改结果')

test2 = r_sort_orders()
test2.to_excel("Data_Order.xlsx", sheet_name='sheet2', index=False)
df2 = pd.read_excel("Data_Order.xlsx", sheet_name='sheet2')
unique_value2 = df2['起始城市'].nunique()
print(unique_value2)
unique_count2 = df2['起始城市'].unique()
print(unique_count2)
# ['兰州市' '哈尔滨市' '沈阳市' '成都市' '乌鲁木齐市' '昆明市' '惠州市' '杭州市' '海口市' '乌兰察布市' '济南市'
#  '福州市' '武汉市' '合肥市' '东莞市' '南京市' '南宁市' '泉州市' '厦门市' '深圳市' '南昌市' '洛阳市' '石家庄市'
#  '太原市' '北京市' '广州市' '嘉兴市' '呼和浩特市' '上海市' '天津市']

